import numpy
# Reading an excel file using Python 
import xlrd 
import xlwt
import xlsxwriter
import openpyxl
  
# Workbook is created 
#wb = Workbook() 
# Give the location of the file 
loc = ( r"C:\Users\USER\Desktop\1.xlsx")
loc1=(r"C:\Users\USER\Desktop\try1.xlsx")
# To open Workbook 
wb =openpyxl.load_workbook(loc)
wb2=openpyxl.load_workbook(loc1)
sheet=wb.get_sheet_by_name('Sheet1')
ws=wb.worksheets[0]
sheet2=wb2.get_sheet_by_name('Sheet2')
# For row 0 and column 0 
rows=235
columns=5
listab=[]
for i in range(1,rows+1):
    listab.append([])
if(ws.cell(row=i,column=4)=='F'):        
    for r in range(1,rows+1):
        for c in range(1,columns+1):
                e=sheet.cell(row=r,column=c)
                listab[r-1].append(e.value)

for r in range(1,rows+1):
    for c in range(1,columns+1):
        j=sheet2.cell(row=r,column=c)
        j.value=listab[r-1]
wb2.save('try1.xlsx')
